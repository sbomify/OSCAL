#!/usr/bin/env python3
"""
Generate an OSCAL 1.1.2 Catalog from the Cyber Essentials 2026 (Danzell v16)
comparison spreadsheet, validated via oscal-pydantic.

Focuses on the Danzell (left-hand / current) column set.

Usage (from repo root):
    uv run --with 'oscal-pydantic-v2,openpyxl' python3 tools/generate_oscal.py
"""

import json
import os
import re
import uuid as uuid_mod
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from oscal_pydantic import catalog as cat_mod
from oscal_pydantic import document as doc_mod
from oscal_pydantic.core import common, properties

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
CE_NS = "https://iasme.co.uk/ns/cyber-essentials"
OSCAL_NS = "http://csrc.nist.gov/ns/oscal"

# Deterministic UUIDs
NS_UUID = uuid_mod.UUID("d7e3f1a0-2b4c-4e6f-8a0b-1c2d3e4f5a6b")


def make_uuid(name: str) -> uuid_mod.UUID:
    return uuid_mod.uuid5(NS_UUID, name)


# Questions that trigger automatic assessment failure
AUTO_FAIL_QUESTIONS = {
    "A6.4", "A6.5",       # patching (OS/firmware + software)
    "A7.16", "A7.17",     # MFA for cloud admins + users
}

# ---------------------------------------------------------------------------
# Section definitions
# ---------------------------------------------------------------------------
SECTIONS = OrderedDict([
    ("A1", {
        "id": "ce-a1-organisation",
        "title": "Organisation",
        "class": "organisational",
        "subsections": {
            "Your Organisation": "ce-a1-org-details",
            "Certificates": "ce-a1-certificates",
            "Application": "ce-a1-application",
        },
    }),
    ("A2", {
        "id": "ce-a2-scope",
        "title": "Scope of Assessment",
        "class": "organisational",
        "subsections": {
            "Scope of Assessment": "ce-a2-scope-details",
        },
    }),
    ("A3", {
        "id": "ce-a3-insurance",
        "title": "Cyber Insurance",
        "class": "organisational",
        "subsections": {},
    }),
    ("A4", {
        "id": "ce-a4-firewalls",
        "title": "Firewalls",
        "class": "technical-control",
        "subsections": {},
    }),
    ("A5", {
        "id": "ce-a5-secure-config",
        "title": "Secure Configuration",
        "class": "technical-control",
        "subsections": {
            "Device Unlocking Method": "ce-a5-device-unlock",
        },
    }),
    ("A6", {
        "id": "ce-a6-update-mgmt",
        "title": "Security Update Management",
        "class": "technical-control",
        "subsections": {},
    }),
    ("A7", {
        "id": "ce-a7-access-control",
        "title": "User Access Control",
        "class": "technical-control",
        "subsections": {
            "Administrative Accounts": "ce-a7-admin-accounts",
            "Password-Based Authentication": "ce-a7-password-auth",
        },
    }),
    ("A8", {
        "id": "ce-a8-malware",
        "title": "Malware Protection",
        "class": "technical-control",
        "subsections": {},
    }),
])


def section_prefix(qno: str) -> str:
    m = re.match(r"(A\d+)", qno)
    return m.group(1) if m else ""


def slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


# ---------------------------------------------------------------------------
# Parse spreadsheet
# ---------------------------------------------------------------------------

def parse_spreadsheet(path: str):
    wb = openpyxl.load_workbook(path)
    ws = wb["DANZELL-WILLOW COMPARISON"]

    questions = []
    section_descriptions: dict[str, str] = {}
    subsection_descriptions: dict[str, str] = {}

    current_category = ""
    current_subcategory = ""

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
        raw_cat = str(row[1].value or "").strip()
        raw_sub = str(row[2].value or "").strip()
        qno = str(row[3].value or "").strip()
        question = str(row[4].value or "").strip()
        guidance = str(row[5].value or "").strip()
        answer_type = str(row[6].value or "").strip()

        if raw_cat == "Category":
            continue

        if not qno:
            if raw_cat:
                current_category = raw_cat
                current_subcategory = raw_sub or ""
                for prefix, sec in SECTIONS.items():
                    if raw_cat in (sec["title"], "Controls") and (
                        not raw_sub or raw_sub == sec["title"]
                        or raw_sub in sec.get("subsections", {})
                        or raw_cat == sec["title"]
                    ):
                        if raw_sub in sec.get("subsections", {}):
                            sub_id = sec["subsections"][raw_sub]
                            subsection_descriptions[sub_id] = question or ""
                        elif raw_cat == sec["title"] or (
                            raw_cat == "Controls" and raw_sub == sec["title"]
                        ):
                            combined = question
                            if raw_sub and raw_sub not in sec.get("subsections", {}):
                                combined = f"{raw_sub}\n\n{question}" if question else raw_sub
                            section_descriptions[sec["id"]] = combined or ""
            elif raw_sub:
                current_subcategory = raw_sub
                if question:
                    for prefix, sec in SECTIONS.items():
                        if raw_sub in sec.get("subsections", {}):
                            sub_id = sec["subsections"][raw_sub]
                            subsection_descriptions[sub_id] = question
            continue

        if raw_cat:
            current_category = raw_cat
        if raw_sub:
            current_subcategory = raw_sub

        questions.append({
            "category": current_category,
            "subcategory": current_subcategory,
            "number": qno,
            "question": question,
            "guidance": guidance,
            "answer_type": answer_type,
        })

    return questions, section_descriptions, subsection_descriptions


# ---------------------------------------------------------------------------
# Extract CE Requirements
# ---------------------------------------------------------------------------

CE_REQ_PATTERN = re.compile(
    r"CE Requirement:\s*(.*?)(?=\n\n|\nFurther guidance|\nPlease note|\nGuidance |$)",
    re.DOTALL,
)


def extract_ce_requirements(guidance: str) -> list[str]:
    reqs = CE_REQ_PATTERN.findall(guidance)
    return [r.strip() for r in reqs if r.strip()]


# ---------------------------------------------------------------------------
# Build OSCAL objects via oscal-pydantic
# ---------------------------------------------------------------------------

def make_prop(name: str, value: str, ns: str | None = None) -> properties.BaseProperty:
    kwargs = {"name": name, "value": value}
    if ns:
        kwargs["ns"] = ns
    return properties.BaseProperty(**kwargs)


def build_control(item: dict) -> cat_mod.Control:
    qno = item["number"]
    control_id = slug(qno)
    question = item["question"]
    guidance = item["guidance"]
    answer_type = item["answer_type"]

    # Determine control class from section
    prefix = section_prefix(qno)
    sec = SECTIONS.get(prefix, {})
    control_class = sec.get("class", "organisational")

    # Properties
    ctrl_props = [make_prop("label", qno)]
    ctrl_props.append(make_prop("sort-id", control_id))

    if answer_type:
        ctrl_props.append(make_prop("response-type", answer_type, ns=CE_NS))

    if qno in AUTO_FAIL_QUESTIONS:
        ctrl_props.append(make_prop("auto-fail", "true", ns=CE_NS))

    # Parts
    parts: list[cat_mod.BasePart] = []

    # Statement
    parts.append(cat_mod.BasePart(
        id=f"{control_id}_smt",
        name="statement",
        prose=question,
    ))

    # Guidance (with CE requirements as nested items)
    if guidance:
        ce_reqs = extract_ce_requirements(guidance)

        # Build child parts for each CE requirement
        req_items = []
        for i, req in enumerate(ce_reqs, 1):
            req_items.append(cat_mod.BasePart(
                id=f"{control_id}_gdn.req{i}",
                name="item",
                part_class="ce-requirement",
                prose=f"CE Requirement: {req}",
            ))

        guidance_part = cat_mod.BasePart(
            id=f"{control_id}_gdn",
            name="guidance",
            prose=guidance,
            parts=req_items if req_items else None,
        )
        parts.append(guidance_part)

    return cat_mod.Control(
        id=control_id,
        control_class=control_class,
        title=question,
        props=ctrl_props,
        parts=parts,
    )


def build_catalog(questions, section_descriptions, subsection_descriptions) -> doc_mod.Document:
    catalog_uuid = make_uuid("ce-danzell-v16-catalog")

    # Parties
    iasme_uuid = make_uuid("iasme")
    ncsc_uuid = make_uuid("ncsc")

    iasme = common.Party(
        uuid=iasme_uuid,
        type="organization",
        name="IASME Consortium",
        remarks=["IASME manages the Cyber Essentials scheme on behalf of the NCSC."],
    )
    ncsc = common.Party(
        uuid=ncsc_uuid,
        type="organization",
        name="National Cyber Security Centre (NCSC)",
        remarks=["NCSC owns the Cyber Essentials scheme."],
    )

    # Roles
    creator_role = common.Role(id="creator", title="Document Creator")
    assessor_role = common.Role(id="assessor", title="Cyber Essentials Assessor")

    metadata = common.Metadata(
        title="Cyber Essentials Question Set — Danzell (April 2026, Version 16)",
        last_modified=datetime.now(timezone.utc).isoformat(),
        version="16.0",
        oscal_version="1.1.2",
        roles=[creator_role, assessor_role],
        parties=[iasme, ncsc],
        responsible_parties=[
            common.ResponsibleParty(
                role_id="creator",
                party_uuids=[iasme_uuid],
            ),
        ],
        remarks=(
            "This OSCAL catalog represents the Cyber Essentials self-assessment "
            "question set version 16 (code-named 'Danzell'), effective April 2026. "
            "It was machine-generated from the official Danzell-Willow comparison "
            "spreadsheet published by IASME."
        ),
    )

    # Back matter with references
    back_matter = common.BackMatter(
        resources=[
            common.Resource(
                uuid=make_uuid("ncsc-ce-req-v3.3"),
                title="Cyber Essentials: Requirements for IT Infrastructure v3.3",
                rlinks=[common.ResourceLink(
                    href="https://www.ncsc.gov.uk/files/cyber-essentials-requirements-for-it-infrastructure-v3-3.pdf",
                )],
                remarks=(
                    "The NCSC Requirements for IT Infrastructure document defines "
                    "the technical requirements for Cyber Essentials certification. "
                    "Version 3.3 applies to assessment accounts created after 27 April 2026."
                ),
            ),
            common.Resource(
                uuid=make_uuid("iasme-ce-overview"),
                title="IASME Cyber Essentials Scheme Overview",
                rlinks=[common.ResourceLink(
                    href="https://iasme.co.uk/cyber-essentials/",
                )],
                remarks="IASME scheme overview and certification body information.",
            ),
            common.Resource(
                uuid=make_uuid("ncsc-ce-overview"),
                title="NCSC Cyber Essentials Overview",
                rlinks=[common.ResourceLink(
                    href="https://www.ncsc.gov.uk/cyberessentials/overview",
                )],
                remarks="Official NCSC Cyber Essentials scheme page.",
            ),
            common.Resource(
                uuid=make_uuid("danzell-willow-comparison"),
                title="Danzell-Willow Comparison Document",
                remarks=(
                    "Source spreadsheet comparing the Danzell (v16, April 2026) "
                    "and Willow (v15, April 2025) question sets."
                ),
            ),
        ],
    )

    # Bucket questions by section
    buckets: dict[str, OrderedDict[str, list]] = {}
    for prefix in SECTIONS:
        buckets[prefix] = OrderedDict()

    for q in questions:
        prefix = section_prefix(q["number"])
        if prefix not in buckets:
            buckets[prefix] = OrderedDict()
        sub = q["subcategory"] or "_default"
        if sub not in buckets[prefix]:
            buckets[prefix][sub] = []
        buckets[prefix][sub].append(q)

    # Build groups
    groups = []
    for prefix, sec_info in SECTIONS.items():
        sec_id = sec_info["id"]

        group_props = [
            make_prop("label", prefix),
        ]

        group_parts = None
        if sec_id in section_descriptions and section_descriptions[sec_id]:
            group_parts = [cat_mod.BasePart(
                id=f"{sec_id}-overview",
                name="overview",
                prose=section_descriptions[sec_id],
            )]

        sub_groups_def = sec_info.get("subsections", {})
        sub_items = buckets.get(prefix, {})

        controls = []
        child_groups = []

        if sub_groups_def and len(sub_items) > 1:
            for sub_name, items in sub_items.items():
                if sub_name in sub_groups_def:
                    sub_id = sub_groups_def[sub_name]
                    sg_parts = None
                    if sub_id in subsection_descriptions and subsection_descriptions[sub_id]:
                        sg_parts = [cat_mod.BasePart(
                            id=f"{sub_id}-overview",
                            name="overview",
                            prose=subsection_descriptions[sub_id],
                        )]
                    sg = cat_mod.Group(
                        id=sub_id,
                        title=sub_name,
                        controls=[build_control(i) for i in items],
                        parts=sg_parts,
                    )
                    child_groups.append(sg)
                else:
                    controls.extend(build_control(i) for i in items)
        else:
            all_items = [i for items in sub_items.values() for i in items]
            if all_items:
                controls = [build_control(i) for i in all_items]

        group = cat_mod.Group(
            id=sec_id,
            group_class=sec_info.get("class"),
            title=sec_info["title"],
            props=group_props,
            parts=group_parts,
            controls=controls if controls else None,
            groups=child_groups if child_groups else None,
        )
        groups.append(group)

    oscal_catalog = cat_mod.Catalog(
        uuid=catalog_uuid,
        metadata=metadata,
        groups=groups,
        back_matter=back_matter,
    )

    return doc_mod.Document(catalog=oscal_catalog)


# ---------------------------------------------------------------------------
# Post-processing: fix oscal-pydantic output to match NIST JSON schema
# ---------------------------------------------------------------------------

def _fixup_schema_compliance(data: dict):
    """Fix known divergences between oscal-pydantic output and the NIST schema."""

    # Fix 1: Party.remarks — oscal-pydantic emits list[str], schema expects string
    for party in data.get("catalog", {}).get("metadata", {}).get("parties", []):
        if isinstance(party.get("remarks"), list):
            party["remarks"] = "\n".join(party["remarks"])

    # Fix 2: Prop values containing newlines — schema pattern is ^\S(.*\S)?$
    # which does not allow embedded newlines. Collapse to single line.
    def fix_props(obj):
        for prop in obj.get("props", []):
            val = prop.get("value", "")
            if "\n" in val:
                prop["value"] = " ".join(line.strip() for line in val.splitlines() if line.strip())
        for ctrl in obj.get("controls", []):
            fix_props(ctrl)
        for grp in obj.get("groups", []):
            fix_props(grp)

    for grp in data.get("catalog", {}).get("groups", []):
        fix_props(grp)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    repo_root = Path(__file__).resolve().parent.parent
    source_path = repo_root / "catalogs" / "cyber-essentials" / "danzell-v16" / "source.xlsx"
    questions, sec_descs, sub_descs = parse_spreadsheet(str(source_path))
    doc = build_catalog(questions, sec_descs, sub_descs)

    # Serialize using oscal-pydantic's built-in serializer (by_alias, exclude_none)
    json_str = doc.model_dump_json()

    # Post-process to fix oscal-pydantic quirks vs NIST schema:
    # 1. Party.remarks is list[str] in pydantic model but string in OSCAL schema
    # 2. Prop values with newlines violate the ^\S(.*\S)?$ pattern
    data = json.loads(json_str)
    _fixup_schema_compliance(data)

    output_path = str(repo_root / "catalogs" / "cyber-essentials" / "danzell-v16" / "catalog.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

    print(f"Generated {output_path}")

    # Stats
    total = 0
    req_count = 0
    auto_fail_count = 0

    def count_controls(obj):
        nonlocal total, req_count, auto_fail_count
        for c in obj.get("controls", []):
            total += 1
            for p in c.get("parts", []):
                if p.get("name") == "guidance":
                    for child in p.get("parts", []):
                        if child.get("class") == "ce-requirement":
                            req_count += 1
            for prop in c.get("props", []):
                if prop.get("name") == "auto-fail":
                    auto_fail_count += 1
        for g in obj.get("groups", []):
            count_controls(g)

    for g in data["catalog"]["groups"]:
        count_controls(g)

    print(f"  Total controls: {total}")
    print(f"  CE Requirements extracted: {req_count}")
    print(f"  Auto-fail controls flagged: {auto_fail_count}")

    from collections import Counter
    sections = Counter(section_prefix(q["number"]) for q in questions)
    for sec, info in SECTIONS.items():
        count = sections.get(sec, 0)
        print(f"    {sec} ({info['title']}): {count} controls")


if __name__ == "__main__":
    main()
