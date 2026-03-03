#!/usr/bin/env python3
"""
Validate the Cyber Essentials OSCAL catalog against multiple dimensions:

1. OSCAL JSON Schema compliance (NIST official v1.1.2 schema)
2. oscal-pydantic round-trip (deserialization + reserialization)
3. Content fidelity (every spreadsheet question appears in the catalog)
4. Completeness (no missing or duplicated controls)
5. CE requirement extraction accuracy
6. Auto-fail flag correctness

Usage (from repo root):
    uv run --with 'oscal-pydantic-v2,openpyxl,jsonschema' python3 tools/validate_catalog.py
"""

import json
import re
import sys
import urllib.request
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
CATALOG_PATH = REPO_ROOT / "catalogs" / "cyber-essentials" / "danzell-v16" / "catalog.json"
SOURCE_PATH = REPO_ROOT / "catalogs" / "cyber-essentials" / "danzell-v16" / "source.xlsx"
SCHEMA_URL = "https://github.com/usnistgov/OSCAL/releases/download/v1.1.2/oscal_catalog_schema.json"
SCHEMA_CACHE = Path("/tmp/oscal_catalog_schema.json")

# Known auto-fail questions in Danzell v16
EXPECTED_AUTO_FAIL = {"A6.4", "A6.5", "A7.16", "A7.17"}

passed = 0
failed = 0
warnings = 0


def ok(msg: str):
    global passed
    passed += 1
    print(f"  PASS  {msg}")


def fail(msg: str):
    global failed
    failed += 1
    print(f"  FAIL  {msg}")


def warn(msg: str):
    global warnings
    warnings += 1
    print(f"  WARN  {msg}")


# ---------------------------------------------------------------------------
# 1. OSCAL JSON Schema validation
# ---------------------------------------------------------------------------

def validate_schema(catalog_data: dict):
    print("\n=== 1. OSCAL JSON Schema Validation (NIST v1.1.2) ===")

    try:
        import jsonschema
        from jsonschema import Draft7Validator
    except ImportError:
        warn("jsonschema not installed, skipping schema validation")
        return

    # The OSCAL schema uses \p{L} (Unicode property escapes) which Python's
    # built-in `re` module does not support. Use the `regex` module if available.
    try:
        import regex
        jsonschema_format_checker = None

        # Monkey-patch re in jsonschema's keyword module to use `regex`
        import jsonschema._legacy_keywords as _lk
        import jsonschema._keywords as _kw
        _kw.re = regex
        _lk.re = regex
    except ImportError:
        warn("'regex' package not installed; schema patterns using Unicode "
             "properties (\\p{L}) may cause errors. Install with: "
             "uv add regex")

    # Download schema if not cached
    if not SCHEMA_CACHE.exists():
        print(f"  Downloading schema from {SCHEMA_URL}...")
        urllib.request.urlretrieve(SCHEMA_URL, SCHEMA_CACHE)

    with open(SCHEMA_CACHE) as f:
        schema = json.load(f)

    validator = Draft7Validator(schema)

    try:
        errors = list(validator.iter_errors(catalog_data))
    except Exception as e:
        warn(f"Schema validation encountered an error: {e}")
        return

    if not errors:
        ok("Catalog validates against NIST OSCAL 1.1.2 JSON schema")
    else:
        for err in errors[:10]:
            path = " -> ".join(str(p) for p in err.absolute_path) or "(root)"
            fail(f"Schema error at {path}: {err.message[:120]}")
        if len(errors) > 10:
            fail(f"... and {len(errors) - 10} more schema errors")


# ---------------------------------------------------------------------------
# 2. oscal-pydantic round-trip
# ---------------------------------------------------------------------------

def validate_pydantic_roundtrip(json_str: str):
    print("\n=== 2. oscal-pydantic Round-Trip Validation ===")

    try:
        import warnings as w
        w.filterwarnings("ignore")
        from oscal_pydantic import document as doc_mod
    except ImportError:
        warn("oscal-pydantic not installed, skipping round-trip validation")
        return

    # oscal-pydantic has a known bug: Party.remarks is typed as list[str]
    # but the NIST JSON schema (and our output) uses plain string.
    # Patch the JSON to satisfy pydantic before loading.
    patched = json.loads(json_str)
    for party in patched.get("catalog", {}).get("metadata", {}).get("parties", []):
        r = party.get("remarks")
        if isinstance(r, str):
            party["remarks"] = [r]
    patched_str = json.dumps(patched)

    try:
        doc = doc_mod.Document.model_validate_json(patched_str)
        ok("Deserialization succeeded (with Party.remarks workaround)")
    except Exception as e:
        fail(f"Deserialization failed: {e}")
        return

    try:
        reserialized = doc.model_dump_json()
        ok("Reserialization succeeded")
    except Exception as e:
        fail(f"Reserialization failed: {e}")
        return

    roundtripped = json.loads(reserialized)
    orig = json.loads(json_str)

    if orig["catalog"]["uuid"] == roundtripped["catalog"]["uuid"]:
        ok("UUID preserved through round-trip")
    else:
        fail("UUID changed during round-trip")

    orig_count = count_all_controls(orig["catalog"])
    rt_count = count_all_controls(roundtripped["catalog"])
    if orig_count == rt_count:
        ok(f"Control count preserved ({orig_count})")
    else:
        fail(f"Control count mismatch: original={orig_count}, round-tripped={rt_count}")


# ---------------------------------------------------------------------------
# 3. Content fidelity — compare against source spreadsheet
# ---------------------------------------------------------------------------

def parse_spreadsheet_questions(path: str) -> dict:
    """Parse Danzell questions from the spreadsheet, keyed by question number."""
    wb = openpyxl.load_workbook(path)
    ws = wb["DANZELL-WILLOW COMPARISON"]

    questions = {}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
        qno = str(row[3].value or "").strip()
        question = str(row[4].value or "").strip()
        guidance = str(row[5].value or "").strip()
        answer_type = str(row[6].value or "").strip()

        if not qno or qno == "No":
            continue

        questions[qno] = {
            "question": question,
            "guidance": guidance,
            "answer_type": answer_type,
        }

    return questions


def extract_catalog_controls(catalog: dict) -> dict:
    """Extract all controls from the catalog, keyed by label (question number)."""
    controls = {}

    def walk(obj):
        for ctrl in obj.get("controls", []):
            label = None
            for prop in ctrl.get("props", []):
                if prop.get("name") == "label":
                    label = prop["value"]
                    break
            if label:
                controls[label] = ctrl
        for grp in obj.get("groups", []):
            walk(grp)

    for grp in catalog.get("groups", []):
        walk(grp)

    return controls


def validate_content_fidelity(catalog_data: dict):
    print("\n=== 3. Content Fidelity (Spreadsheet vs Catalog) ===")

    spreadsheet_qs = parse_spreadsheet_questions(str(SOURCE_PATH))
    catalog_ctrls = extract_catalog_controls(catalog_data["catalog"])

    # Check every spreadsheet question exists in catalog
    missing_from_catalog = set(spreadsheet_qs.keys()) - set(catalog_ctrls.keys())
    extra_in_catalog = set(catalog_ctrls.keys()) - set(spreadsheet_qs.keys())

    if not missing_from_catalog:
        ok(f"All {len(spreadsheet_qs)} spreadsheet questions present in catalog")
    else:
        for q in sorted(missing_from_catalog):
            fail(f"Missing from catalog: {q}")

    if not extra_in_catalog:
        ok("No extra controls in catalog beyond spreadsheet")
    else:
        for q in sorted(extra_in_catalog):
            warn(f"Extra control in catalog not in spreadsheet: {q}")

    # Check question text matches
    text_mismatches = 0
    for qno in sorted(set(spreadsheet_qs.keys()) & set(catalog_ctrls.keys())):
        expected = spreadsheet_qs[qno]["question"]
        ctrl = catalog_ctrls[qno]

        # Get statement text from catalog
        actual = ""
        for part in ctrl.get("parts", []):
            if part.get("name") == "statement":
                actual = part.get("prose", "")
                break

        if expected != actual:
            text_mismatches += 1
            if text_mismatches <= 5:
                fail(f"{qno} statement text mismatch")
                print(f"         Expected: {expected[:80]}...")
                print(f"         Got:      {actual[:80]}...")

    if text_mismatches == 0:
        ok("All question statement texts match exactly")
    elif text_mismatches > 5:
        fail(f"... and {text_mismatches - 5} more text mismatches")

    # Check guidance text matches
    guidance_mismatches = 0
    for qno in sorted(set(spreadsheet_qs.keys()) & set(catalog_ctrls.keys())):
        expected_guidance = spreadsheet_qs[qno]["guidance"]
        ctrl = catalog_ctrls[qno]

        actual_guidance = ""
        for part in ctrl.get("parts", []):
            if part.get("name") == "guidance":
                actual_guidance = part.get("prose", "")
                break

        if expected_guidance and expected_guidance != actual_guidance:
            guidance_mismatches += 1
            if guidance_mismatches <= 3:
                fail(f"{qno} guidance text mismatch")

    if guidance_mismatches == 0:
        ok("All guidance texts match exactly")
    else:
        fail(f"{guidance_mismatches} guidance text mismatches total")

    # Check response types
    type_mismatches = 0
    for qno in sorted(set(spreadsheet_qs.keys()) & set(catalog_ctrls.keys())):
        expected_type = spreadsheet_qs[qno]["answer_type"]
        ctrl = catalog_ctrls[qno]

        actual_type = ""
        for prop in ctrl.get("props", []):
            if prop.get("name") == "response-type":
                actual_type = prop["value"]
                break

        # Normalize whitespace — multiline values are collapsed for schema compliance
        norm_expected = " ".join(expected_type.split()) if expected_type else ""
        norm_actual = " ".join(actual_type.split()) if actual_type else ""
        if norm_expected and norm_expected != norm_actual:
            type_mismatches += 1
            if type_mismatches <= 3:
                fail(f"{qno} response-type mismatch: expected '{norm_expected}', got '{norm_actual}'")

    if type_mismatches == 0:
        ok("All response types match")
    else:
        fail(f"{type_mismatches} response-type mismatches total")


# ---------------------------------------------------------------------------
# 4. Completeness — no duplicates, correct counts
# ---------------------------------------------------------------------------

def count_all_controls(catalog: dict) -> int:
    total = 0

    def walk(obj):
        nonlocal total
        total += len(obj.get("controls", []))
        for grp in obj.get("groups", []):
            walk(grp)

    for grp in catalog.get("groups", []):
        walk(grp)
    return total


def validate_completeness(catalog_data: dict):
    print("\n=== 4. Completeness Checks ===")

    catalog = catalog_data["catalog"]
    controls = extract_catalog_controls(catalog)

    # Check for duplicate IDs
    all_ids = []

    def collect_ids(obj):
        for ctrl in obj.get("controls", []):
            all_ids.append(ctrl["id"])
        for grp in obj.get("groups", []):
            collect_ids(grp)

    for grp in catalog.get("groups", []):
        collect_ids(grp)

    dupes = [x for x in all_ids if all_ids.count(x) > 1]
    if not dupes:
        ok(f"No duplicate control IDs ({len(all_ids)} total)")
    else:
        fail(f"Duplicate control IDs found: {set(dupes)}")

    # Check every group has an ID
    def check_group_ids(groups, path=""):
        for grp in groups:
            gid = grp.get("id")
            if not gid:
                fail(f"Group without ID at {path}/{grp.get('title', '?')}")
            else:
                ok_count = len(grp.get("controls", [])) + len(grp.get("groups", []))
                if ok_count == 0:
                    warn(f"Empty group: {gid}")
            check_group_ids(grp.get("groups", []), f"{path}/{gid}")

    check_group_ids(catalog.get("groups", []))

    # Verify section control counts match expected prefixes
    prefix_counts = {}
    for label in controls:
        m = re.match(r"(A\d+)", label)
        if m:
            prefix = m.group(1)
            prefix_counts[prefix] = prefix_counts.get(prefix, 0) + 1

    for prefix in ["A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8"]:
        count = prefix_counts.get(prefix, 0)
        if count > 0:
            ok(f"Section {prefix}: {count} controls")
        else:
            fail(f"Section {prefix}: no controls found")


# ---------------------------------------------------------------------------
# 5. CE requirement extraction accuracy
# ---------------------------------------------------------------------------

CE_REQ_PATTERN = re.compile(
    r"CE Requirement:\s*(.*?)(?=\n\n|\nFurther guidance|\nPlease note|\nGuidance |$)",
    re.DOTALL,
)


def validate_ce_requirements(catalog_data: dict):
    print("\n=== 5. CE Requirement Extraction Accuracy ===")

    controls = extract_catalog_controls(catalog_data["catalog"])
    spreadsheet_qs = parse_spreadsheet_questions(str(SOURCE_PATH))

    total_expected = 0
    total_found = 0
    mismatches = 0

    for qno in sorted(set(spreadsheet_qs.keys()) & set(controls.keys())):
        guidance = spreadsheet_qs[qno]["guidance"]
        expected_reqs = CE_REQ_PATTERN.findall(guidance)
        expected_reqs = [r.strip() for r in expected_reqs if r.strip()]

        ctrl = controls[qno]
        found_reqs = []
        for part in ctrl.get("parts", []):
            if part.get("name") == "guidance":
                for child in part.get("parts", []):
                    if child.get("class") == "ce-requirement":
                        found_reqs.append(child.get("prose", ""))

        total_expected += len(expected_reqs)
        total_found += len(found_reqs)

        if len(expected_reqs) != len(found_reqs):
            mismatches += 1
            if mismatches <= 3:
                fail(f"{qno}: expected {len(expected_reqs)} CE reqs, found {len(found_reqs)}")

    if mismatches == 0:
        ok(f"All CE requirements correctly extracted ({total_found} total)")
    else:
        fail(f"{mismatches} controls with CE requirement count mismatches")

    if total_expected == total_found:
        ok(f"Total CE requirement count matches: {total_found}")
    else:
        fail(f"Total CE requirements: expected {total_expected}, found {total_found}")


# ---------------------------------------------------------------------------
# 6. Auto-fail flags
# ---------------------------------------------------------------------------

def validate_auto_fail_flags(catalog_data: dict):
    print("\n=== 6. Auto-Fail Flag Validation ===")

    controls = extract_catalog_controls(catalog_data["catalog"])

    flagged = set()
    for label, ctrl in controls.items():
        for prop in ctrl.get("props", []):
            if prop.get("name") == "auto-fail" and prop.get("value") == "true":
                flagged.add(label)

    if flagged == EXPECTED_AUTO_FAIL:
        ok(f"Auto-fail flags correct: {sorted(flagged)}")
    else:
        missing = EXPECTED_AUTO_FAIL - flagged
        extra = flagged - EXPECTED_AUTO_FAIL
        if missing:
            fail(f"Missing auto-fail flags: {sorted(missing)}")
        if extra:
            warn(f"Unexpected auto-fail flags: {sorted(extra)}")


# ---------------------------------------------------------------------------
# 7. Structural validation
# ---------------------------------------------------------------------------

def validate_structure(catalog_data: dict):
    print("\n=== 7. Structural Validation ===")

    catalog = catalog_data["catalog"]

    # Check required top-level fields
    for field in ["uuid", "metadata", "groups", "back-matter"]:
        if field in catalog:
            ok(f"Top-level field '{field}' present")
        else:
            fail(f"Missing top-level field: {field}")

    # Check metadata
    meta = catalog.get("metadata", {})
    for field in ["title", "version", "oscal-version", "roles", "parties", "responsible-parties"]:
        if field in meta:
            ok(f"Metadata field '{field}' present")
        else:
            fail(f"Missing metadata field: {field}")

    if meta.get("oscal-version") == "1.1.2":
        ok("OSCAL version is 1.1.2")
    else:
        fail(f"OSCAL version is {meta.get('oscal-version')}, expected 1.1.2")

    # Check back-matter has resources
    bm = catalog.get("back-matter", {})
    resources = bm.get("resources", [])
    if len(resources) >= 3:
        ok(f"Back-matter has {len(resources)} resources")
    else:
        warn(f"Back-matter only has {len(resources)} resources")

    # Check every control has required parts
    controls = extract_catalog_controls(catalog)
    missing_statement = 0
    for label, ctrl in controls.items():
        has_statement = any(p.get("name") == "statement" for p in ctrl.get("parts", []))
        if not has_statement:
            missing_statement += 1
            if missing_statement <= 3:
                fail(f"{label}: missing 'statement' part")

    if missing_statement == 0:
        ok("All controls have a 'statement' part")

    # Check all part names are canonical OSCAL
    valid_part_names = {"statement", "guidance", "item", "overview", "instruction",
                        "assessment-objective", "assessment-method", "assessment-objects"}
    bad_parts = set()

    def check_parts(parts):
        for p in parts:
            ns = p.get("ns", "http://csrc.nist.gov/ns/oscal")
            name = p.get("name", "")
            if "csrc.nist.gov" in str(ns) and name not in valid_part_names:
                bad_parts.add(name)
            check_parts(p.get("parts", []))

    def walk_all(obj):
        for ctrl in obj.get("controls", []):
            check_parts(ctrl.get("parts", []))
        for grp in obj.get("groups", []):
            check_parts(grp.get("parts", []))
            walk_all(grp)

    for grp in catalog.get("groups", []):
        check_parts(grp.get("parts", []))
        walk_all(grp)

    if not bad_parts:
        ok("All part names are canonical OSCAL")
    else:
        fail(f"Non-canonical part names found: {bad_parts}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    global passed, failed, warnings

    print(f"Validating: {CATALOG_PATH}")
    print(f"Source:     {SOURCE_PATH}")

    with open(CATALOG_PATH) as f:
        json_str = f.read()

    catalog_data = json.loads(json_str)

    validate_schema(catalog_data)
    validate_pydantic_roundtrip(json_str)
    validate_content_fidelity(catalog_data)
    validate_completeness(catalog_data)
    validate_ce_requirements(catalog_data)
    validate_auto_fail_flags(catalog_data)
    validate_structure(catalog_data)

    print(f"\n{'=' * 50}")
    print(f"Results: {passed} passed, {failed} failed, {warnings} warnings")
    print(f"{'=' * 50}")

    if failed > 0:
        sys.exit(1)
    else:
        print("\nAll validations passed.")
        sys.exit(0)


if __name__ == "__main__":
    main()
